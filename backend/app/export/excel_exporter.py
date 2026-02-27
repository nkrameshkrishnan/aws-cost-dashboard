"""
Excel report generator using openpyxl for multi-sheet FinOps audit reports.
"""
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference

from app.schemas.audit import FullAuditResults


class ExcelReportGenerator:
    """Generate Excel workbooks with multiple sheets for FinOps audits."""

    def __init__(self):
        """Initialize Excel generator."""
        # Define color scheme
        self.colors = {
            'header': 'F3F4F6',
            'critical': 'DC2626',
            'warning': 'F59E0B',
            'success': '10B981',
            'info': '3B82F6'
        }

    def generate_audit_report(self, audit_results: FullAuditResults) -> bytes:
        """
        Generate a multi-sheet Excel audit report.

        Args:
            audit_results: Full audit results

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create Summary sheet
        self._create_summary_sheet(wb, audit_results)

        # Create EC2 sheet
        if audit_results.ec2_audit and (audit_results.ec2_audit.idle_instances or audit_results.ec2_audit.stopped_instances):
            self._create_ec2_sheet(wb, audit_results.ec2_audit)

        # Create EBS sheet
        if audit_results.ebs_audit and (audit_results.ebs_audit.unattached_volumes or audit_results.ebs_audit.old_snapshots):
            self._create_ebs_sheet(wb, audit_results.ebs_audit)

        # Create RDS sheet
        if audit_results.rds_audit and audit_results.rds_audit.idle_instances:
            self._create_rds_sheet(wb, audit_results.rds_audit)

        # Create Lambda sheet
        if audit_results.lambda_audit and (audit_results.lambda_audit.unused_functions or audit_results.lambda_audit.over_provisioned_functions):
            self._create_lambda_sheet(wb, audit_results.lambda_audit)

        # Create S3 sheet
        if audit_results.s3_audit and (audit_results.s3_audit.buckets_without_lifecycle or audit_results.s3_audit.incomplete_multipart_uploads):
            self._create_s3_sheet(wb, audit_results.s3_audit)

        # Create NAT Gateway sheet
        if audit_results.nat_gateway_audit and (audit_results.nat_gateway_audit.idle_gateways or audit_results.nat_gateway_audit.unused_gateways):
            self._create_nat_gateway_sheet(wb, audit_results.nat_gateway_audit)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        return excel_bytes

    def _create_summary_sheet(self, wb: Workbook, audit_results: FullAuditResults):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = 'FinOps Audit Report - Executive Summary'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
        ws.merge_cells('A1:D1')

        # Account Info
        ws['A3'] = 'Account Name:'
        ws['B3'] = audit_results.account_name
        ws['A4'] = 'Audit Date:'
        ws['B4'] = audit_results.audit_timestamp.strftime('%Y-%m-%d %H:%M:%S')

        # Key Metrics
        ws['A6'] = 'Key Metrics'
        ws['A6'].font = Font(size=14, bold=True)
        ws.merge_cells('A6:D6')

        headers = ['Metric', 'Value']
        ws.append([])
        ws.append(headers)

        # Style headers
        for col in range(1, 3):
            cell = ws.cell(row=8, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # Metrics data
        metrics = [
            ('Total Findings', audit_results.summary.total_findings),
            ('Potential Monthly Savings', f"${audit_results.summary.total_potential_savings:,.2f}"),
            ('Critical Severity', audit_results.summary.findings_by_severity.get('critical', 0)),
            ('High Severity', audit_results.summary.findings_by_severity.get('high', 0)),
            ('Medium Severity', audit_results.summary.findings_by_severity.get('medium', 0)),
            ('Low Severity', audit_results.summary.findings_by_severity.get('low', 0)),
        ]

        for metric, value in metrics:
            ws.append([metric, value])

        # Top Opportunities
        if audit_results.summary.top_opportunities:
            ws.append([])
            ws.append(['Top Savings Opportunities'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=14, bold=True)
            ws.merge_cells(f'A{last_row}:D{last_row}')

            ws.append([])
            for idx, opportunity in enumerate(audit_results.summary.top_opportunities, 1):
                ws.append([f"{idx}.", opportunity])

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _create_ec2_sheet(self, wb: Workbook, ec2_audit):
        """Create EC2 findings sheet."""
        ws = wb.create_sheet("EC2 Findings")

        # Title
        ws['A1'] = 'EC2 Findings'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
        ws.merge_cells('A1:G1')

        # Summary
        ws['A3'] = 'Summary'
        ws['A3'].font = Font(size=12, bold=True)
        ws.append([])
        ws.append(['Idle Instances:', len(ec2_audit.idle_instances)])
        ws.append(['Stopped Instances:', len(ec2_audit.stopped_instances)])
        ws.append(['Total Potential Savings:', f"${ec2_audit.total_potential_savings:,.2f}/month"])

        # Idle Instances Table
        if ec2_audit.idle_instances:
            ws.append([])
            ws.append(['Idle Instances (Low CPU Utilization)'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Instance ID', 'Instance Type', 'Region', 'Avg CPU %', 'Launch Time', 'Est. Monthly Cost', 'Potential Savings']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for instance in ec2_audit.idle_instances:
                ws.append([
                    instance.instance_id,
                    instance.instance_type,
                    instance.region,
                    f"{instance.avg_cpu_utilization:.2f}%",
                    instance.launch_time.strftime('%Y-%m-%d') if instance.launch_time else 'N/A',
                    f"${instance.estimated_monthly_cost:.2f}",
                    f"${instance.potential_monthly_savings:.2f}"
                ])

            # Add total row
            total_savings = sum(i.potential_monthly_savings for i in ec2_audit.idle_instances)
            ws.append(['', '', '', '', '', 'TOTAL:', f"${total_savings:,.2f}"])
            total_row = ws.max_row
            ws[f'F{total_row}'].font = Font(bold=True)
            ws[f'G{total_row}'].font = Font(bold=True)

        # Stopped Instances Table
        if ec2_audit.stopped_instances:
            ws.append([])
            ws.append([])
            ws.append(['Stopped Instances (Still Incurring EBS Costs)'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Instance ID', 'Instance Type', 'Region', 'Days Stopped', 'Est. EBS Cost']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for instance in ec2_audit.stopped_instances:
                ws.append([
                    instance.instance_id,
                    instance.instance_type,
                    instance.region,
                    instance.days_stopped,
                    f"${instance.estimated_ebs_cost:.2f}"
                ])

        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_ebs_sheet(self, wb: Workbook, ebs_audit):
        """Create EBS findings sheet."""
        ws = wb.create_sheet("EBS Findings")

        # Title
        ws['A1'] = 'EBS Findings'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
        ws.merge_cells('A1:F1')

        # Summary
        ws['A3'] = 'Summary'
        ws['A3'].font = Font(size=12, bold=True)
        ws.append([])
        ws.append(['Unattached Volumes:', len(ebs_audit.unattached_volumes)])
        ws.append(['Old Snapshots:', len(ebs_audit.old_snapshots)])
        ws.append(['Total Potential Savings:', f"${ebs_audit.total_potential_savings:,.2f}/month"])

        # Unattached Volumes Table
        if ebs_audit.unattached_volumes:
            ws.append([])
            ws.append(['Unattached EBS Volumes'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Volume ID', 'Size (GB)', 'Volume Type', 'Region', 'Days Unattached', 'Est. Monthly Cost']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for volume in ebs_audit.unattached_volumes:
                ws.append([
                    volume.volume_id,
                    volume.size_gb,
                    volume.volume_type,
                    volume.region,
                    volume.days_unattached,
                    f"${volume.estimated_monthly_cost:.2f}"
                ])

            # Add total row
            total_cost = sum(v.estimated_monthly_cost for v in ebs_audit.unattached_volumes)
            ws.append(['', '', '', '', 'TOTAL:', f"${total_cost:,.2f}"])
            total_row = ws.max_row
            ws[f'E{total_row}'].font = Font(bold=True)
            ws[f'F{total_row}'].font = Font(bold=True)

        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_rds_sheet(self, wb: Workbook, rds_audit):
        """Create RDS findings sheet."""
        ws = wb.create_sheet("RDS Findings")

        # Title
        ws['A1'] = 'RDS Findings'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
        ws.merge_cells('A1:H1')

        # Summary
        ws['A3'] = 'Summary'
        ws['A3'].font = Font(size=12, bold=True)
        ws.append([])
        ws.append(['Idle Instances:', len(rds_audit.idle_instances)])
        ws.append(['Total Potential Savings:', f"${rds_audit.total_potential_savings:,.2f}/month"])

        # Idle Instances Table
        if rds_audit.idle_instances:
            ws.append([])
            ws.append(['Idle RDS Instances'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Instance ID', 'Instance Class', 'Engine', 'Region', 'Avg CPU %', 'Avg Connections', 'Est. Monthly Cost', 'Potential Savings']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for instance in rds_audit.idle_instances:
                ws.append([
                    instance.db_instance_id,
                    instance.db_instance_class,
                    instance.engine,
                    instance.region,
                    f"{instance.avg_cpu_utilization:.2f}%",
                    f"{instance.avg_connections:.0f}",
                    f"${instance.estimated_monthly_cost:.2f}",
                    f"${instance.potential_monthly_savings:.2f}"
                ])

            # Add total row
            total_savings = sum(i.potential_monthly_savings for i in rds_audit.idle_instances)
            ws.append(['', '', '', '', '', '', 'TOTAL:', f"${total_savings:,.2f}"])
            total_row = ws.max_row
            ws[f'G{total_row}'].font = Font(bold=True)
            ws[f'H{total_row}'].font = Font(bold=True)

        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 16

    def _create_lambda_sheet(self, wb: Workbook, lambda_audit):
        """Create Lambda findings sheet."""
        ws = wb.create_sheet("Lambda Findings")

        # Title
        ws['A1'] = 'Lambda Findings'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
        ws.merge_cells('A1:G1')

        # Summary
        ws['A3'] = 'Summary'
        ws['A3'].font = Font(size=12, bold=True)
        ws.append([])
        ws.append(['Unused Functions:', len(lambda_audit.unused_functions)])
        ws.append(['Over-Provisioned Functions:', len(lambda_audit.over_provisioned_functions)])
        ws.append(['Total Potential Savings:', f"${lambda_audit.total_potential_savings:,.2f}/month"])

        # Unused Functions Table
        if lambda_audit.unused_functions:
            ws.append([])
            ws.append(['Unused Lambda Functions (No Invocations)'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Function Name', 'Region', 'Runtime', 'Memory (MB)', 'Days Unused', 'Est. Monthly Cost']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for func in lambda_audit.unused_functions:
                ws.append([
                    func.function_name,
                    func.region,
                    func.runtime,
                    func.memory_mb,
                    func.days_since_invocation,
                    f"${func.estimated_monthly_cost:.2f}"
                ])

            # Add total row
            total_cost = sum(f.estimated_monthly_cost for f in lambda_audit.unused_functions)
            ws.append(['', '', '', '', 'TOTAL:', f"${total_cost:,.2f}"])
            total_row = ws.max_row
            ws[f'E{total_row}'].font = Font(bold=True)
            ws[f'F{total_row}'].font = Font(bold=True)

        # Over-Provisioned Functions Table
        if lambda_audit.over_provisioned_functions:
            ws.append([])
            ws.append(['Over-Provisioned Lambda Functions'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Function Name', 'Region', 'Configured Memory (MB)', 'Avg Used Memory (MB)', 'Utilization %', 'Est. Monthly Cost', 'Potential Savings']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for func in lambda_audit.over_provisioned_functions:
                ws.append([
                    func.function_name,
                    func.region,
                    func.configured_memory_mb,
                    f"{func.avg_memory_used_mb:.0f}",
                    f"{func.memory_utilization_percent:.1f}%",
                    f"${func.estimated_monthly_cost:.2f}",
                    f"${func.potential_monthly_savings:.2f}"
                ])

            # Add total row
            total_savings = sum(f.potential_monthly_savings for f in lambda_audit.over_provisioned_functions)
            ws.append(['', '', '', '', '', 'TOTAL:', f"${total_savings:,.2f}"])
            total_row = ws.max_row
            ws[f'F{total_row}'].font = Font(bold=True)
            ws[f'G{total_row}'].font = Font(bold=True)

        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_s3_sheet(self, wb: Workbook, s3_audit):
        """Create S3 findings sheet."""
        ws = wb.create_sheet("S3 Findings")

        # Title
        ws['A1'] = 'S3 Findings'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
        ws.merge_cells('A1:E1')

        # Summary
        ws['A3'] = 'Summary'
        ws['A3'].font = Font(size=12, bold=True)
        ws.append([])
        ws.append(['Buckets Without Lifecycle:', len(s3_audit.buckets_without_lifecycle)])
        ws.append(['Incomplete Multipart Uploads:', len(s3_audit.incomplete_multipart_uploads)])
        ws.append(['Total Potential Savings:', f"${s3_audit.total_potential_savings:,.2f}/month"])

        # Buckets Without Lifecycle Table
        if s3_audit.buckets_without_lifecycle:
            ws.append([])
            ws.append(['Buckets Without Lifecycle Policies'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Bucket Name', 'Region', 'Size (GB)', 'Object Count', 'Est. Monthly Cost', 'Potential Savings']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for bucket in s3_audit.buckets_without_lifecycle:
                ws.append([
                    bucket.bucket_name,
                    bucket.region,
                    f"{bucket.total_size_gb:.2f}",
                    bucket.object_count,
                    f"${bucket.estimated_monthly_cost:.2f}",
                    f"${bucket.potential_monthly_savings:.2f}"
                ])

            # Add total row
            total_savings = sum(b.potential_monthly_savings for b in s3_audit.buckets_without_lifecycle)
            ws.append(['', '', '', '', 'TOTAL:', f"${total_savings:,.2f}"])
            total_row = ws.max_row
            ws[f'E{total_row}'].font = Font(bold=True)
            ws[f'F{total_row}'].font = Font(bold=True)

        # Incomplete Multipart Uploads Table
        if s3_audit.incomplete_multipart_uploads:
            ws.append([])
            ws.append([])
            ws.append(['Incomplete Multipart Uploads'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['Bucket Name', 'Key', 'Days Old', 'Size (GB)', 'Est. Monthly Cost']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for upload in s3_audit.incomplete_multipart_uploads:
                ws.append([
                    upload.bucket_name,
                    upload.key,
                    upload.days_old,
                    f"{upload.estimated_size_gb:.2f}",
                    f"${upload.estimated_monthly_cost:.2f}"
                ])

            # Add total row
            total_cost = sum(u.estimated_monthly_cost for u in s3_audit.incomplete_multipart_uploads)
            ws.append(['', '', '', 'TOTAL:', f"${total_cost:,.2f}"])
            total_row = ws.max_row
            ws[f'D{total_row}'].font = Font(bold=True)
            ws[f'E{total_row}'].font = Font(bold=True)

        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22

    def _create_nat_gateway_sheet(self, wb: Workbook, nat_gateway_audit):
        """Create NAT Gateway findings sheet."""
        ws = wb.create_sheet("NAT Gateway Findings")

        # Title
        ws['A1'] = 'NAT Gateway Findings'
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
        ws.merge_cells('A1:F1')

        # Summary
        ws['A3'] = 'Summary'
        ws['A3'].font = Font(size=12, bold=True)
        ws.append([])
        ws.append(['Idle Gateways:', len(nat_gateway_audit.idle_gateways)])
        ws.append(['Unused Gateways:', len(nat_gateway_audit.unused_gateways)])
        ws.append(['Total Potential Savings:', f"${nat_gateway_audit.total_potential_savings:,.2f}/month"])

        # Idle NAT Gateways Table
        if nat_gateway_audit.idle_gateways:
            ws.append([])
            ws.append(['Idle NAT Gateways (Low Traffic)'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['NAT Gateway ID', 'Region', 'VPC ID', 'Avg GB Out/Day', 'Est. Monthly Cost', 'Potential Savings']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for gateway in nat_gateway_audit.idle_gateways:
                ws.append([
                    gateway.nat_gateway_id,
                    gateway.region,
                    gateway.vpc_id,
                    f"{gateway.avg_gb_out_per_day:.2f}",
                    f"${gateway.estimated_monthly_cost:.2f}",
                    f"${gateway.potential_monthly_savings:.2f}"
                ])

            # Add total row
            total_savings = sum(g.potential_monthly_savings for g in nat_gateway_audit.idle_gateways)
            ws.append(['', '', '', '', 'TOTAL:', f"${total_savings:,.2f}"])
            total_row = ws.max_row
            ws[f'E{total_row}'].font = Font(bold=True)
            ws[f'F{total_row}'].font = Font(bold=True)

        # Unused NAT Gateways Table
        if nat_gateway_audit.unused_gateways:
            ws.append([])
            ws.append([])
            ws.append(['Unused NAT Gateways (No Traffic)'])
            last_row = ws.max_row
            ws[f'A{last_row}'].font = Font(size=12, bold=True)

            # Headers
            headers = ['NAT Gateway ID', 'Region', 'VPC ID', 'Days Active', 'Est. Monthly Cost']
            ws.append(headers)
            header_row = ws.max_row

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                cell.alignment = Alignment(horizontal='left', vertical='center')

            # Data
            for gateway in nat_gateway_audit.unused_gateways:
                ws.append([
                    gateway.nat_gateway_id,
                    gateway.region,
                    gateway.vpc_id,
                    gateway.days_active,
                    f"${gateway.estimated_monthly_cost:.2f}"
                ])

            # Add total row
            total_cost = sum(g.estimated_monthly_cost for g in nat_gateway_audit.unused_gateways)
            ws.append(['', '', '', 'TOTAL:', f"${total_cost:,.2f}"])
            total_row = ws.max_row
            ws[f'D{total_row}'].font = Font(bold=True)
            ws[f'E{total_row}'].font = Font(bold=True)

        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def generate_cost_report(self, cost_data: Dict[str, Any], profile_name: str) -> bytes:
        """
        Generate Excel report for cost data (Dashboard).

        Args:
            cost_data: Dictionary containing daily_costs and service_breakdown
            profile_name: AWS profile name

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Summary sheet
        ws_summary = wb.create_sheet('Summary')
        ws_summary.append(['AWS Cost Report'])
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.append(['Profile:', profile_name])
        ws_summary.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append([])

        # Daily costs sheet
        if 'daily_costs' in cost_data and cost_data['daily_costs']:
            ws_daily = wb.create_sheet('Daily Costs')
            ws_daily.append(['Date', 'Cost (USD)'])

            # Header styling
            for col in range(1, 3):
                cell = ws_daily.cell(row=1, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data
            total_cost = 0
            for item in cost_data['daily_costs']:
                cost = float(item.get('cost', 0))
                total_cost += cost
                ws_daily.append([item.get('date', ''), f"${cost:.2f}"])

            # Total row
            ws_daily.append(['TOTAL', f"${total_cost:,.2f}"])
            total_row = ws_daily.max_row
            ws_daily[f'A{total_row}'].font = Font(bold=True)
            ws_daily[f'B{total_row}'].font = Font(bold=True)

            # Column widths
            ws_daily.column_dimensions['A'].width = 15
            ws_daily.column_dimensions['B'].width = 15

        # Service breakdown sheet
        if 'service_breakdown' in cost_data and cost_data['service_breakdown']:
            ws_services = wb.create_sheet('Service Breakdown')
            ws_services.append(['Service', 'Cost (USD)', 'Percentage'])

            # Header styling
            for col in range(1, 4):
                cell = ws_services.cell(row=1, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data
            total_cost = sum(float(item.get('cost', 0)) for item in cost_data['service_breakdown'])
            for item in cost_data['service_breakdown']:
                cost = float(item.get('cost', 0))
                percentage = (cost / total_cost * 100) if total_cost > 0 else 0
                ws_services.append([
                    item.get('service', 'Unknown'),
                    f"${cost:.2f}",
                    f"{percentage:.1f}%"
                ])

            # Column widths
            ws_services.column_dimensions['A'].width = 30
            ws_services.column_dimensions['B'].width = 15
            ws_services.column_dimensions['C'].width = 12

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_forecast_report(self, forecast_data: Dict[str, Any], profile_name: str) -> bytes:
        """
        Generate Excel report for forecast data (Analytics).

        Args:
            forecast_data: Dictionary containing forecast predictions
            profile_name: AWS profile name

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Forecast sheet
        ws = wb.create_sheet('Cost Forecast')
        ws.append(['AWS Cost Forecast'])
        ws['A1'].font = Font(size=16, bold=True)
        ws.append(['Profile:', profile_name])
        ws.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([])

        # Summary
        ws.append(['Forecast Period Start:', forecast_data.get('forecast_period_start', 'N/A')])
        ws.append(['Forecast Period End:', forecast_data.get('forecast_period_end', 'N/A')])
        ws.append(['Forecasted Cost:', f"${forecast_data.get('forecasted_cost', 0):,.2f}"])
        ws.append([])

        # Predictions if available
        if 'predictions' in forecast_data and forecast_data['predictions']:
            ws.append(['Date', 'Forecasted Cost (USD)'])

            # Header styling
            for col in range(1, 3):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data
            for item in forecast_data['predictions']:
                ws.append([item.get('date', ''), f"${item.get('cost', 0):.2f}"])

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_rightsizing_report(self, rightsizing_data: Dict[str, Any], profile_name: str) -> bytes:
        """
        Generate Excel report for right-sizing recommendations.

        Args:
            rightsizing_data: Dictionary containing recommendations
            profile_name: AWS profile name

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Summary sheet
        ws = wb.create_sheet('Right-Sizing Recommendations')
        ws.append(['AWS Right-Sizing Report'])
        ws['A1'].font = Font(size=16, bold=True)
        ws.append(['Profile:', profile_name])
        ws.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([])

        ws.append(['Total Recommendations:', rightsizing_data.get('total_recommendations', 0)])
        ws.append(['Total Monthly Savings:', f"${rightsizing_data.get('total_monthly_savings', 0):,.2f}"])
        ws.append([])

        # Recommendations if available
        if 'recommendations' in rightsizing_data and rightsizing_data['recommendations']:
            ws.append(['Resource ID', 'Resource Type', 'Current Type', 'Recommended Type', 'Monthly Savings'])

            # Header styling
            for col in range(1, 6):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data
            for item in rightsizing_data['recommendations']:
                ws.append([
                    item.get('resource_id', 'N/A'),
                    item.get('resource_type', 'N/A'),
                    item.get('current_type', 'N/A'),
                    item.get('recommended_type', 'N/A'),
                    f"${item.get('monthly_savings', 0):.2f}"
                ])

            # Column widths
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
